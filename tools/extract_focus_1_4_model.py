"""Extract the machine-readable FOCUS 1.4 data model from the committed Excel.

Dev-time tool (requires ``openpyxl``). Reads the FinOps "FOCUS 1.4 Data Model"
workbook (download it from https://focus.finops.org — it is not committed to
this repository) plus the ServiceSubcategory supplement, and writes the
deterministic, committed ``src/focus_data_toolkit/model/focus_1_4_model.json``
that the FOCUS 1.4 validator consumes.

The committed JSON is the artifact of record; neither the validator nor the test
suite needs openpyxl. Re-run after the workbook changes:

    python tools/extract_focus_1_4_model.py /path/to/focus_1_4_data_model.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

_HERE = Path(__file__).resolve()
_REPO = _HERE.parents[1]
_XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else _REPO / "focus_1_4_data_model.xlsx"
_MODEL_DIR = _REPO / "src" / "focus_data_toolkit" / "model"
_SUPPLEMENT = _MODEL_DIR / "focus_1_4_servicesubcategory.json"
_OUT = _MODEL_DIR / "focus_1_4_model.json"

# Datasets we model (the four FOCUS 1.4 datasets). Order is stable for output.
_DATASET_ORDER = ("Cost and Usage", "Billing Period", "Contract Commitment", "Invoice Detail")


def _cell(row: tuple, idx: dict, name: str) -> str:
    j = idx.get(name)
    if j is None or j >= len(row) or row[j] is None:
        return ""
    return str(row[j]).strip()


def _as_bool(value: str) -> bool:
    return value.strip().lower() == "true"


def _parse_allowed(raw: str, column_id: str, value_format: str, supplement: list[str]):
    """Return (allowed_values, numeric_range) parsed from the 'Allowed Values' cell."""
    raw = (raw or "").strip()
    if column_id == "ServiceSubcategory":
        return sorted(supplement), None  # supplement maps subcategory -> parent category
    if not raw or raw.lower() == "n/a":
        return None, None
    rng = re.fullmatch(r"([0-9]*\.?[0-9]+)\s*to\s*([0-9]*\.?[0-9]+)", raw)
    if rng:
        return None, [float(rng.group(1)), float(rng.group(2))]
    if raw.lower().startswith("see ") or raw.startswith("["):
        return None, None  # external reference or expected-format hint, not an enum
    if value_format == "Allowed Values":
        return [v.strip() for v in raw.split(",") if v.strip()], None
    return None, None


def build_model() -> dict:
    import openpyxl  # local import: dev-time dependency only

    supplement = json.loads(_SUPPLEMENT.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)

    # Dataset metadata (short id + feature level) from the 'Datasets' sheet.
    ds_rows = list(wb["Datasets"].iter_rows(values_only=True))
    ds_idx = {h: j for j, h in enumerate(ds_rows[0]) if h}
    ds_meta: dict[str, dict] = {}
    for row in ds_rows[1:]:
        name = _cell(row, ds_idx, "Dataset")
        if name in _DATASET_ORDER:
            ds_meta[name] = {
                "dataset_id": _cell(row, ds_idx, "Dataset ID"),
                "short_id": _cell(row, ds_idx, "Dataset Short ID"),
                "dataset_type": _cell(row, ds_idx, "Dataset Type"),
                "feature_level": _cell(row, ds_idx, "Feature Level"),
            }

    # Columns from the 'Columns' sheet.
    col_rows = list(wb["Columns"].iter_rows(values_only=True))
    cidx = {h: j for j, h in enumerate(col_rows[0]) if h}
    datasets: dict[str, dict] = {
        name: {**ds_meta.get(name, {}), "columns": {}} for name in _DATASET_ORDER
    }
    for row in col_rows[1:]:
        dataset = _cell(row, cidx, "Dataset")
        column_id = _cell(row, cidx, "Column ID")
        if dataset not in datasets or not column_id:
            continue
        if _as_bool(_cell(row, cidx, "Removed")):
            continue  # removed columns are not part of FOCUS 1.4
        value_format = _cell(row, cidx, "Value Format")
        allowed, numeric_range = _parse_allowed(
            _cell(row, cidx, "Allowed Values"), column_id, value_format, supplement
        )
        datasets[dataset]["columns"][column_id] = {
            "display_name": _cell(row, cidx, "Display Name"),
            "category": _cell(row, cidx, "Category"),
            "version": _cell(row, cidx, "Ver"),
            "feature_level": _cell(row, cidx, "Feature Level"),
            "condition": _cell(row, cidx, "Condition"),
            "deprecated": _as_bool(_cell(row, cidx, "Deprecated")),
            "allows_nulls": _as_bool(_cell(row, cidx, "Allows Nulls")),
            "data_type": _cell(row, cidx, "Data Type"),
            "value_format": value_format,
            "allowed_values": allowed,
            "numeric_range": numeric_range,
        }

    return {
        "focus_version": "1.4",
        "source": (
            "FinOps FOCUS 1.4 Data Model workbook (https://focus.finops.org) + "
            "FOCUS spec v1.4 ServiceSubcategory list"
        ),
        "service_subcategory_parents": supplement,
        "datasets": datasets,
    }


def main() -> int:
    model = build_model()
    _OUT.write_text(json.dumps(model, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    counts = {ds: len(v["columns"]) for ds, v in model["datasets"].items()}
    print(f"wrote {_OUT}")
    print("column counts:", counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
